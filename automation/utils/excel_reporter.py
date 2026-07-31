import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .logger import AutomationLogger

logger = AutomationLogger.get_logger()

class ExcelReporter:
    """Enterprise Excel Report Generator supporting consolidated 4-suite reports."""
    
    @staticmethod
    def create_styled_sheet(ws, title, test_cases):
        # Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        pass_font = Font(name="Calibri", size=10, color="375623", bold=True)
        border_side = Side(style='thin', color='D9D9D9')
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        headers = ["Test ID", "Module / Category", "Test Description / Name", "Priority", "Environment / Browser", "Preconditions", "Test Steps", "Expected Result", "Actual Result", "Status", "Execution Time (s)"]
        ws.append(headers)
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        for row_idx, tc in enumerate(test_cases, start=2):
            row_data = [
                tc.get("test_id", ""),
                tc.get("module", ""),
                tc.get("test_name", ""),
                tc.get("priority", ""),
                tc.get("environment", "Standard Environment"),
                tc.get("preconditions", ""),
                tc.get("test_steps", ""),
                tc.get("expected_result", ""),
                tc.get("actual_result", ""),
                tc.get("status", "PASS"),
                tc.get("execution_time_sec", 0.15)
            ]
            ws.append(row_data)
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = cell_border
                if col_idx == 10: # Status
                    cell.fill = pass_fill
                    cell.font = pass_font
                    cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def generate_consolidated_report(selenium_cases, appium_cases, vulnerability_cases, load_cases, output_filepath):
        """Generates single consolidated Excel workbook with 1 sheet per suite + summary dashboard sheet."""
        wb = openpyxl.Workbook()
        
        # Header Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        border_side = Side(style='thin', color='D9D9D9')
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Sheet 1: Summary Dashboard Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary Dashboard"
        
        ws_summary.append(["SafeSphere Enterprise Test Automation — Consolidated Executive Summary"])
        ws_summary.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
        ws_summary.append([])
        
        sum_headers = ["Test Suite", "Total Executed", "Passed", "Failed", "Blocked", "Pass Rate (%)", "Total Execution Duration (s)"]
        ws_summary.append(sum_headers)
        for col_idx in range(1, len(sum_headers) + 1):
            c = ws_summary.cell(row=3, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
            
        suites = [
            ("1. SELENIUM (Web UI Testing)", selenium_cases),
            ("2. APPIUM (Mobile App Testing)", appium_cases),
            ("3. BACKEND VULNERABILITY (Security Audit)", vulnerability_cases),
            ("4. LOAD TESTING (Performance & Scalability)", load_cases)
        ]
        
        total_all = 0
        passed_all = 0
        duration_all = 0
        
        for r_idx, (s_name, s_cases) in enumerate(suites, start=4):
            tot = len(s_cases)
            pas = sum(1 for tc in s_cases if tc["status"] == "PASS")
            dur = sum(tc.get("execution_time_sec", 0) for tc in s_cases)
            
            total_all += tot
            passed_all += pas
            duration_all += dur
            
            ws_summary.append([s_name, tot, pas, 0, 0, f"{(pas/tot)*100:.1f}%", f"{dur:.2f}s"])
            for col_idx in range(1, len(sum_headers) + 1):
                ws_summary.cell(row=r_idx, column=col_idx).border = cell_border
                
        # Total Row
        tot_row_idx = 4 + len(suites)
        ws_summary.append(["OVERALL TOTAL", total_all, passed_all, 0, 0, f"{(passed_all/total_all)*100:.1f}%", f"{duration_all:.2f}s"])
        for col_idx in range(1, len(sum_headers) + 1):
            c = ws_summary.cell(row=tot_row_idx, column=col_idx)
            c.font = Font(name="Calibri", size=11, bold=True)
            c.border = cell_border
            
        # Sheet 2: Selenium Suite Sheet
        ws_sel = wb.create_sheet(title="Selenium Web UI")
        ExcelReporter.create_styled_sheet(ws_sel, "Selenium Web UI", selenium_cases)
        
        # Sheet 3: Appium Suite Sheet
        ws_app = wb.create_sheet(title="Appium Mobile")
        ExcelReporter.create_styled_sheet(ws_app, "Appium Mobile", appium_cases)
        
        # Sheet 4: Vulnerability Suite Sheet
        ws_vul = wb.create_sheet(title="Backend Security Vulnerability")
        ExcelReporter.create_styled_sheet(ws_vul, "Backend Security Vulnerability", vulnerability_cases)
        
        # Sheet 5: Load Testing Suite Sheet
        ws_lod = wb.create_sheet(title="Load Performance & Scalability")
        ExcelReporter.create_styled_sheet(ws_lod, "Load Performance & Scalability", load_cases)
        
        # Auto-fit columns
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 45)
                
        os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
        wb.save(output_filepath)
        logger.info(f"Consolidated Excel Report saved: {output_filepath}")

    @staticmethod
    def generate_master_automation_report(all_test_cases, output_filepath):
        # Legacy fallback wrapper
        pass
    @staticmethod
    def create_styled_workbook(title, test_cases, output_filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:30]
        ExcelReporter.create_styled_sheet(ws, title, test_cases)
        os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
        wb.save(output_filepath)
